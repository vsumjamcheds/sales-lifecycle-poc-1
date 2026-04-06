"""
Export HCP lists for RepGA / RepNJ / RepFL to an Excel workbook (one sheet per rep + combined).

Usage (from project root, venv active):
  python -m src.scripts.export_rep_hcp_excel

Output: exports/rep_hcp_lists.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.database.models import HCP
from src.database.session import SessionLocal

TERRITORY_TO_REP = {"GA": "RepGA", "NJ": "RepNJ", "FL": "RepFL"}
HEADERS = ["rep_code", "territory_code", "hcp_id", "display_name", "specialty"]


def run(out_path: Path | None = None) -> Path:
    if out_path is None:
        from src.config import settings

        out_path = settings.project_root / "exports" / "rep_hcp_lists.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    db = SessionLocal()
    try:
        hcps = db.query(HCP).order_by(HCP.territory_code, HCP.display_name).all()
    finally:
        db.close()

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_all = wb.create_sheet("All_Reps", 0)
    ws_all.append(HEADERS)
    for h in hcps:
        rep = TERRITORY_TO_REP.get(h.territory_code, h.territory_code)
        ws_all.append([rep, h.territory_code, h.id, h.display_name, h.specialty])

    for terr, rep_code in [("GA", "RepGA"), ("NJ", "RepNJ"), ("FL", "RepFL")]:
        ws = wb.create_sheet(rep_code)
        ws.append(["hcp_id", "display_name", "specialty", "territory_code"])
        for h in hcps:
            if h.territory_code == terr:
                ws.append([h.id, h.display_name, h.specialty, h.territory_code])
        _autosize_columns(ws)

    _autosize_columns(ws_all)
    wb.save(out_path)
    print(f"Wrote {out_path} ({len(hcps)} HCPs).")
    return out_path


def _autosize_columns(ws, max_width: int = 48) -> None:
    for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        max_len = 0
        for cell in column:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


if __name__ == "__main__":
    run()
